
import pandas as pd
from typing import Dict, List, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import shutil
import os
import sys

class ExcelHandler:
    """
    Класс для обработки Excel файлов
    """

    @staticmethod
    def load_excel_data(file_path: str) -> Dict[str, pd.DataFrame]:
        """
        Загружает все листы из Excel файла

        Args:
            file_path (str): Путь к Excel файлу

        Returns:
            Dict[str, pd.DataFrame]: Словарь с данными всех листов
        """
        try:
            # Загружаем все листы из Excel файла
            excel_data = pd.read_excel(file_path, sheet_name=None)

            if not excel_data:
                print("No data found in the Excel file")
                return {}

            print(f"Successfully loaded {len(excel_data)} sheets from {file_path}")
            return excel_data

        except FileNotFoundError:
            print(f"File not found: {file_path}")
            return {}
        except Exception as e:
            print(f"Error loading Excel file {file_path}: {e}")
            return {}

    @staticmethod
    def get_column_max_lengths(df: pd.DataFrame) -> Dict[str, int]:
        """
        Вычисляет длину самого длинного значения в каждой колонке DataFrame

        Args:
            df (pd.DataFrame): DataFrame для анализа

        Returns:
            Dict[str, int]: Словарь {имя_колонки: максимальная_длина}
        """
        max_lengths = {}

        for column in df.columns:
            # Преобразуем все значения в строки и находим максимальную длину
            max_length = df[column].astype(str).apply(len).max()
            max_lengths[column] = max_length

        return max_lengths

    @staticmethod
    def get_all_sheets_max_lengths(excel_data: Dict[str, pd.DataFrame]) -> Dict[str, Dict[str, int]]:
        """
        Вычисляет максимальные длины значений для всех колонок во всех листах

        Args:
            excel_data (Dict[str, pd.DataFrame]): Данные Excel (словарь DataFrame'ов)

        Returns:
            Dict[str, Dict[str, int]]: Словарь {имя_листа: {имя_колонки: максимальная_длина}}
        """
        all_max_lengths = {}

        for sheet_name, df in excel_data.items():
            all_max_lengths[sheet_name] = ExcelHandler.get_column_max_lengths(df)

        return all_max_lengths

# Функции для удобного импорта
def load_excel_data(file_path: str) -> Dict[str, pd.DataFrame]:
    """
    Функция-обертка для удобного импорта
    """
    return ExcelHandler.load_excel_data(file_path)

def get_column_max_lengths(df: pd.DataFrame) -> Dict[str, int]:
    """
    Вычисляет длину самого длинного значения в каждой колонке
    """
    return ExcelHandler.get_column_max_lengths(df)

def get_all_sheets_max_lengths(excel_data: Dict[str, pd.DataFrame]) -> Dict[str, Dict[str, int]]:
    """
    Вычисляет максимальные длины значений для всех колонок во всех листах
    """
    return ExcelHandler.get_all_sheets_max_lengths(excel_data)

def excel_processing(file_path: str, sheet_data: Dict[str, pd.DataFrame], 
                    selected_rows: Dict[str, List[int]], save_path: str, 
                    columns_to_keep: List[int]):
    """
    Обработка выделенных данных из Excel файла и сохранение результата

    Args:
        file_path (str): Путь к исходному файлу
        sheet_data (Dict[str, pd.DataFrame]): Данные всех листов
        selected_rows (Dict[str, List[int]]): Выделенные строки по листам
        save_path (str): Путь для сохранения результата
        columns_to_keep (List[int]): Список столбцов (1..25) для сохранения
    """

    # ---- Справочные соответствия колонок старого файла ----
    Old_Columns: Tuple[Tuple[object, object], ...] = (
        ("ISIN", 0),
        ("Ticker &", "Exchange"), 
        ("Ccy", 0),
        ("Cpn", "(%)"),
        ("Name", 0),
        ("Sector", 0),
        ("Industry", 0), 
        ("Maturity", "(1. call date)"),
        ("Price", 0),
        ("Perf", "YTD %"),
        ("Mk-Cap", "mia"),
        ("YTM", "MID"),
        ("Share", "classes"),
        ("ER/MF", 0),
        ("Rating", "Mood"),
        ("Rating", "S&P"),
        ("Rating", "Fitch"),
        ("Size", "mio"),
        ("Z-", "Spread"),
        ("ASW", "spread"),
        ("Min", "piece"),
        ("Min", "incr"),
        ("Mkt of", "Issue"),
        ("Notes", 0),
        ("Added", "on")
    )

    Old_SetColumns = {
        ("ISIN", "0"):1, 
        ("Ticker &", "Exchange"):2, 
        ("Ccy", "0"):3, 
        ("Cpn", "(%)"):4, 
        ("0", "(%)"):4, 
        ("Name", "1"):5, 
        ("Sector", "0"):6, 
        ("Industry", "0"):7, 
        ("Maturity", "(1. call date)"):8, 
        ("Price", "MID"):9, 
        ("Price", "1"):9, 
        ("Perf", "YTD %"):10,
        ("Mk-Cap", "mia"):11, 
        ("YTM", "MID"):12, 
        ("Share class", "0"):13, 
        ("Share", "class"):13, 
        ("ER/MF", "0"):14, 
        ("Rating", "Moody"):15, 
        ("Rating", "S&P"):16, 
        ("Rating", "Fitch"):17, 
        ("Size", "mio"):18, 
        ("Z-", "spread"):19, 
        ("ASW", "spread"):20, 
        ("Min", "piece"):21, 
        ("Min", "incr"):22, 
        (0, "Mkt of Issue"):23, 
        ("Notes", "0"):24, 
        (0, "Notes"):24, 
        ("Added on", "0"):25, 
        ("Added", "on"):25
    }

    # ---- Преобразование columns_to_keep и построение справочников ----
    # Исправление "tuple index out of range": columns_to_keep даны как 1..25,
    # а Old_Columns индексируется с 0, поэтому вычитаем 1.
    Columns: List[Tuple[object, object]] = []
    SetColumns: Dict[Tuple[object, object], int] = {}

    for chosen_column in columns_to_keep:
        if 1 <= chosen_column <= len(Old_Columns):
            Columns.append(Old_Columns[chosen_column - 1])  # <-- фикс off-by-one
            # набираем соответствующие пары заголовков для сопоставления
            for title_for_old_columns, idx in Old_SetColumns.items():
                if idx == chosen_column:
                    SetColumns[title_for_old_columns] = chosen_column

    Chosen_assets: Dict[str, set] = {"Без названия": set()}
    Used_positions: set = set()

    # ---- Подготовка файла результата ----
    file_name = os.path.basename(file_path)
    name_without_ext = os.path.splitext(file_name)[0]
    result_file_name = f"{name_without_ext}_result.xlsx"
    result_path = os.path.join(save_path, result_file_name)

    # Копируем cleaned.xlsx в результат

    if getattr(sys, 'frozen', False):
        # If the application is run as a bundle, the PyInstaller bootloader
        # extends the sys module by a flag frozen=True and sets the app 
        # path into variable _MEIPASS'.
        current_dir = sys._MEIPASS
        print("asddd", sys._MEIPASS)
    else:
        current_dir = os.path.dirname(os.path.abspath(__file__))
    cleaned_path = os.path.join(current_dir, "cleaned.xlsx")
    shutil.copy2(cleaned_path, result_path)

    # Загружаем исходный файл и целевой файл
    source_wb = load_workbook(file_path, read_only=False, data_only=True)
    target_wb = load_workbook(result_path)
    target_ws = target_wb.active

    # ---- Сбор выбранных данных из исходного файла ----
    for source_ws in source_wb.worksheets:
        current_title = "Без названия"
        max_row = source_ws.max_row
        max_column = source_ws.max_column
        dic_to_copy: Dict[int, int] = {}

        # Сопоставляем столбцы по двум строкам заголовка (r2/r3)
        for title_column in range(3, max_column + 1):
            cell1 = source_ws.cell(row=2, column=title_column)
            cell2 = source_ws.cell(row=3, column=title_column)
            value1 = cell1.value if cell1.value is not None else "0"
            value2 = cell2.value if cell2.value is not None else "0"
            title_to_check = (value1, value2)
            title_to_check2 = (value1, "1")

            if title_to_check in SetColumns:
                dic_to_copy[title_column] = SetColumns[title_to_check]
            elif title_to_check2 in SetColumns:
                dic_to_copy[title_column] = SetColumns[title_to_check2]

        # Строки с данными начинаются с 5-й
        for row in source_ws.iter_rows(min_row=5, max_row=max_row, min_col=3, max_col=3):
            cell_C = row[0]
            current_row_idx = cell_C.row
            value_C = cell_C.value
            if not value_C:
                continue

            str_value_C = str(value_C)

            # Если это название группы (не ISIN), обновляем текущий заголовок группы
            if len(str_value_C) != 12:
                if str_value_C not in Chosen_assets:
                    Chosen_assets[str_value_C] = set()
                current_title = str_value_C
            else:
                # Это ISIN/бумага
                if current_row_idx in selected_rows.get(source_ws.title, []):
                    row_cells = next(source_ws.iter_rows(min_row=current_row_idx,
                                                         max_row=current_row_idx))

                    # 25 позиций, как в старом формате
                    list_of_values = [None] * 25

                    for source_col, target_pos in dic_to_copy.items():
                        if source_col - 1 < len(row_cells) and 1 <= target_pos <= 25:
                            list_of_values[target_pos - 1] = row_cells[source_col - 1].value
                            if row_cells[source_col - 1].value:
                                Used_positions.add(target_pos)

                    Chosen_assets[current_title].add(tuple(list_of_values))

    source_wb.close()

    # ---- Подготовка структуры столбцов для вывода ----
    Empty_columns = set(range(1, 26)) - Used_positions
    visible_positions = sorted(pos for pos in range(1, 26) if pos not in Empty_columns)
    visible_count = len(visible_positions)

    # ---- Заголовок портфеля ----
    target_ws.cell(row=1, column=2, value="Balanced Portfolio")
    target_ws.cell(row=1, column=2).alignment = Alignment(horizontal='center', vertical='center')
    target_ws.cell(row=1, column=2).font = Font(size=40, bold=True, color='808080', name='Calabria Light')
    if visible_count >= 2:
        target_ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=visible_count)

    # ---- Заголовки столбцов (2 строки) ----
    current_row = 4
    # Ставим только для реально используемых позиций, в "сжатом" порядке
    for new_idx, old_pos in enumerate(visible_positions, start=1):
        if 1 <= old_pos <= len(Old_Columns):
            top, bottom = Old_Columns[old_pos - 1]
            target_ws.cell(row=2, column=new_idx).value = top if top != 0 else None
            target_ws.cell(row=3, column=new_idx).value = bottom if bottom != 0 else None

    # ---- Базовые стили ----
    FONT_GROUP = Font(bold=True, color='808080', name='Calabria Light')
    FILL_GROUP = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
    BORDER_THIN = Border(left=Side(style=None), right=Side(style=None),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    ALIGN_LEFT = Alignment(horizontal='left', vertical='bottom')

    # ---- Вставка данных ----
    for key, value_set in Chosen_assets.items():
        if not value_set:
            continue

        # Строка группы
        target_ws.cell(row=current_row, column=1, value=key)
        for i in range(1, visible_count + 1):
            c = target_ws.cell(row=current_row, column=i)
            c.font = FONT_GROUP
            c.fill = FILL_GROUP
        current_row += 1

        # Строки значений
        for tuple_item in value_set:
            # tuple_item гарантированно длиной 25
            empty_before = 0
            for col_idx, value in enumerate(tuple_item, start=1):
                if col_idx in Empty_columns:
                    empty_before += 1
                    continue

                new_col = col_idx - empty_before
                if new_col < 1 or new_col > visible_count:
                    continue  # защита от выхода за пределы

                cell_out = target_ws.cell(row=current_row, column=new_col, value=value)
                cell_out.alignment = ALIGN_LEFT
                if isinstance(value, float):
                    cell_out.number_format = '#,##0.0'
            current_row += 1

    # ---- Рамки и авто-ширина ----
    # Рамки
    for i in range(1, visible_count + 1):
        for j in range(4, target_ws.max_row + 1):
            target_ws.cell(row=j, column=i).border = BORDER_THIN

    # Пробегаемся по каждой видимой колонке и считаем максимальную длину текста
    for i, old_pos in enumerate(visible_positions, start=1):
        max_len = 0
        for r in range(1, target_ws.max_row + 1):
            v = target_ws.cell(row=r, column=i).value
            if v is None:
                continue
            length = len(str(v))
            # ограничиваем учётные длины (например, максимум 50 символов)
            if length > 50:
                length = 50
            if length > max_len:
                max_len = length
        # ширина = макс. длина + 2, но не больше 60
        target_ws.column_dimensions[get_column_letter(i)].width = 15

    # ---- Сохранение результата ----
    target_wb.save(result_path)
    target_wb.close()

# Для тестирования модуля
if __name__ == "__main__":
    # Пример использования
    test_xlsx = "test.xlsx"
    if os.path.exists(test_xlsx):
        test_data = load_excel_data(test_xlsx)
        if test_data:
            for sheet_name, df in test_data.items():
                print(f"Sheet: {sheet_name}")
                print(f"Shape: {df.shape}")
                print(f"Columns: {list(df.columns)}")

                # Пример использования новой функциональности
                max_lengths = get_column_max_lengths(df)
                print("Max lengths per column:")
                for col, max_len in max_lengths.items():
                    print(f"  {col}: {max_len}")

                print("-" * 50)

            # Пример для всех листов
            all_max_lengths = get_all_sheets_max_lengths(test_data)
            print("\nMax lengths for all sheets:")
            for sheet_name, lengths in all_max_lengths.items():
                print(f"{sheet_name}: {lengths}")
    else:
        print("test.xlsx not found — skip demo.")
